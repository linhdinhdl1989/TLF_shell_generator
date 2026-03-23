"""Tests for document_parser — text extraction and chunking from file bytes."""
import io
import pytest
from unittest.mock import patch, MagicMock


# ---------------------------------------------------------------------------
# Helpers to build minimal fixture bytes without real files
# ---------------------------------------------------------------------------

def make_docx_bytes(text_paragraphs: list[str]) -> bytes:
    from docx import Document as DocxDocument
    doc = DocxDocument()
    for para in text_paragraphs:
        doc.add_paragraph(para)
    buf = io.BytesIO()
    doc.save(buf)
    buf.seek(0)
    return buf.read()


def make_xlsx_bytes(rows: list[list]) -> bytes:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# parse_document — dispatcher
# ---------------------------------------------------------------------------

def test_parse_document_dispatches_pdf():
    from services.document_parser import parse_document

    fake_page = MagicMock()
    fake_page.extract_text.return_value = "Table 14.1.1 Demographics\n\nSome content here."
    fake_pdf = MagicMock()
    fake_pdf.__enter__ = MagicMock(return_value=fake_pdf)
    fake_pdf.__exit__ = MagicMock(return_value=False)
    fake_pdf.pages = [fake_page]

    with patch("services.document_parser.pdfplumber") as mock_plumber:
        mock_plumber.open.return_value = fake_pdf
        result = parse_document(b"fake-pdf-bytes", "sap.pdf", "sap")

    assert "parsed_content" in result
    assert "chunks_meta" in result
    assert "14.1.1" in result["parsed_content"]


def test_parse_document_dispatches_docx(minimal_docx_bytes):
    from services.document_parser import parse_document

    result = parse_document(minimal_docx_bytes, "sap.docx", "sap")
    assert "parsed_content" in result
    assert "chunks_meta" in result
    assert len(result["parsed_content"]) > 0


def test_parse_document_dispatches_xlsx(minimal_xlsx_bytes):
    from services.document_parser import parse_document

    result = parse_document(minimal_xlsx_bytes, "tlf_list.xlsx", "study_tlf_list")
    assert "parsed_content" in result
    assert len(result["parsed_content"]) > 0


def test_parse_document_unknown_extension_raises():
    from services.document_parser import parse_document, UnsupportedDocumentType

    with pytest.raises(UnsupportedDocumentType):
        parse_document(b"data", "file.xyz", "other")


# ---------------------------------------------------------------------------
# _parse_docx
# ---------------------------------------------------------------------------

def test_parse_docx_extracts_all_paragraphs():
    from services.document_parser import _parse_docx

    content = make_docx_bytes([
        "Statistical Analysis Plan",
        "Table 14.1.1 Demographics",
        "Table 14.2.1 Efficacy",
    ])
    text = _parse_docx(content)
    assert "Statistical Analysis Plan" in text
    assert "14.1.1" in text
    assert "14.2.1" in text


def test_parse_docx_skips_empty_paragraphs():
    from services.document_parser import _parse_docx

    content = make_docx_bytes(["Hello", "", "World"])
    text = _parse_docx(content)
    assert "Hello" in text
    assert "World" in text


# ---------------------------------------------------------------------------
# _parse_xlsx
# ---------------------------------------------------------------------------

def test_parse_xlsx_extracts_cell_values():
    from services.document_parser import _parse_xlsx

    content = make_xlsx_bytes([
        ["Number", "Title", "Section"],
        ["14.1.1", "Demographics", "demographics"],
    ])
    text = _parse_xlsx(content)
    assert "14.1.1" in text
    assert "Demographics" in text


def test_parse_xlsx_multiple_sheets():
    from services.document_parser import _parse_xlsx
    import openpyxl

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["Table 14.1.1", "Demographics"])
    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["Table 14.2.1", "Efficacy"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    text = _parse_xlsx(buf.read())
    assert "14.1.1" in text
    assert "14.2.1" in text


# ---------------------------------------------------------------------------
# _chunk_text
# ---------------------------------------------------------------------------

def test_chunk_text_returns_list_of_dicts():
    from services.document_parser import _chunk_text

    chunks = _chunk_text("Paragraph one.\n\nParagraph two.")
    assert isinstance(chunks, list)
    assert all("chunk_id" in c for c in chunks)
    assert all("section_ref" in c for c in chunks)
    assert all("text" in c for c in chunks)
    assert all("embedding_id" in c for c in chunks)


def test_chunk_text_splits_on_blank_lines():
    from services.document_parser import _chunk_text, _MIN_CHUNK_CHARS

    # Paragraphs must exceed _MIN_CHUNK_CHARS to appear as chunks.
    para = "x" * (_MIN_CHUNK_CHARS + 20)
    text = f"{para}\n\n{para}"
    chunks = _chunk_text(text)
    assert len(chunks) == 2


def test_chunk_text_heading_becomes_section_ref():
    from services.document_parser import _chunk_text

    text = "5.1 Safety Analysis\n\nThis is the content of the safety analysis section."
    chunks = _chunk_text(text)
    content_chunks = [c for c in chunks if "content of the safety" in c["text"]]
    assert len(content_chunks) >= 1
    assert "5.1" in content_chunks[0]["section_ref"]


def test_chunk_text_unique_chunk_ids():
    from services.document_parser import _chunk_text

    text = "\n\n".join([f"Paragraph {i} with sufficient text content." for i in range(5)])
    chunks = _chunk_text(text)
    ids = [c["chunk_id"] for c in chunks]
    assert len(ids) == len(set(ids))


def test_chunk_text_empty_returns_empty():
    from services.document_parser import _chunk_text

    assert _chunk_text("") == []
    assert _chunk_text("   \n\n  ") == []


# ---------------------------------------------------------------------------
# Round-trip: parse_document output feeds extract_tlf_candidates_from_sap
# ---------------------------------------------------------------------------

def test_parsed_docx_feeds_extraction(minimal_docx_bytes):
    """Round-trip: parse a docx, then extract TLF candidates from the text."""
    from services.document_parser import parse_document
    from services.tlf_extraction_service import extract_tlf_candidates_from_sap

    result = parse_document(minimal_docx_bytes, "sap.docx", "sap")
    candidates = extract_tlf_candidates_from_sap(result["parsed_content"])
    # minimal_docx_bytes contains "Table 14.1.1" and "Table 14.2.1"
    numbers = [c["number"] for c in candidates]
    assert "14.1.1" in numbers
    assert "14.2.1" in numbers
